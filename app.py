"""
app.py  —  Kurumsal Coklu Marka Arac Fiyat Scraper Web UI & REST API Sunucusu
=============================================================================
Kullanim:
    python app.py
    (Web tarayicisinda http://localhost:5000 adresine gidin)
"""

from __future__ import annotations

import csv
import io
import os
import sqlite3
import threading
from datetime import datetime
from functools import wraps
from pathlib import Path

from flask import (
    Flask,
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    url_for,
    send_file,
)
from werkzeug.security import check_password_hash, generate_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scrapers.base_scraper import DB_PATH, init_db, log_run, save_records, fmt_price
from scrapers.vw_scraper import VWScraper
from scrapers.skoda_scraper import SkodaScraper
from scrapers.renault_scraper import RenaultScraper
from scrapers.ford_scraper import FordScraper
from scrapers.hyundai_scraper import HyundaiScraper
from scrapers.toyota_scraper import ToyotaScraper
from scrapers.chery_scraper import CheryScraper
from scrapers.dacia_scraper import DaciaScraper
from scrapers.kia_scraper import KiaScraper
from scrapers.fiat_scraper import FiatScraper
from scrapers.peugeot_scraper import PeugeotScraper
from scrapers.opel_scraper import OpelScraper
from scrapers.citroen_scraper import CitroenScraper
from scrapers.jeep_scraper import JeepScraper
from scrapers.alfaromeo_scraper import AlfaRomeoScraper
from scrapers.ds_scraper import DSScraper
from scrapers.maserati_scraper import MaseratiScraper

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "car_price_scraper_secret_key_2026")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

ALL_SCRAPERS = [
    VWScraper(),
    SkodaScraper(),
    RenaultScraper(),
    FordScraper(),
    HyundaiScraper(),
    ToyotaScraper(),
    CheryScraper(),
    DaciaScraper(),
    KiaScraper(),
    FiatScraper(),
    PeugeotScraper(),
    OpelScraper(),
    CitroenScraper(),
    JeepScraper(),
    AlfaRomeoScraper(),
    DSScraper(),
    MaseratiScraper(),
]

SCRAPER_MAP = {s.brand.lower(): s for s in ALL_SCRAPERS}
SCRAPE_STATUS = {"running": False, "message": "Bosta", "progress": 0, "last_run": None}

# Tofaş Grubu Distribütörlük Kapsamı (Stellantis & Tofaş OEM)
OEM_GROUPS = {
    "tofas": [
        "Fiat", "Peugeot", "Opel", "Citroën", "Jeep", "Alfa Romeo", "DS Automobiles", "Maserati"
    ],
}


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    init_db(conn)
    return conn


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "Yetkisiz erisim. Lutfen giris yapin."}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


# --- Auth Rotalari ------------------------------------------------------------

@app.route("/login", methods=["GET", "POST"])
def login():
    conn = get_db()
    users = conn.execute("SELECT * FROM users").fetchall()
    conn.close()

    if not users:
        return redirect(url_for("setup_admin"))

    if session.get("user"):
        return redirect(url_for("index"))

    if request.method == "POST":
        username_input = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        matched_user = None
        for u in users:
            if username_input:
                if u["username"].lower() == username_input.lower() and check_password_hash(u["password_hash"], password):
                    matched_user = u
                    break
            else:
                if check_password_hash(u["password_hash"], password):
                    matched_user = u
                    break

        if matched_user:
            session["user"] = matched_user["username"]
            return redirect(url_for("index"))
        else:
            flash("Hatali kullanici adi veya sifre!", "danger")

    return render_template("login.html")


@app.route("/setup-admin", methods=["GET", "POST"])
def setup_admin():
    conn = get_db()
    admin_user = conn.execute("SELECT * FROM users WHERE username = 'admin'").fetchone()

    if admin_user:
        conn.close()
        return redirect(url_for("login"))

    if request.method == "POST":
        password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")

        if not password or len(password) < 4:
            flash("Sifre en az 4 karakter olmalidir.", "warning")
        elif password != confirm:
            flash("Sifreler eslesmiyor!", "danger")
        else:
            pwd_hash = generate_password_hash(password)
            conn.execute(
                "INSERT INTO users (username, password_hash, created_at) VALUES (?, ?, ?)",
                ("admin", pwd_hash, datetime.now().isoformat()),
            )
            conn.commit()
            conn.close()
            session["user"] = "admin"
            flash("Admin sifreniz basariyla olusturuldu!", "success")
            return redirect(url_for("index"))

    conn.close()
    return render_template("setup.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# --- Web UI Ana Sayfa ---------------------------------------------------------

@app.route("/")
@login_required
def index():
    return render_template("index.html")


# --- API Endpoints -----------------------------------------------------------

@app.route("/api/summary")
@login_required
def api_summary():
    conn = get_db()
    
    last_date_row = conn.execute("SELECT MAX(scraped_date) as max_date FROM prices").fetchone()
    last_date = last_date_row["max_date"] if last_date_row and last_date_row["max_date"] else None

    # Toplam model/varyant sayisi (is_latest = 1)
    total_models = conn.execute(
        "SELECT COUNT(*) as cnt FROM prices WHERE is_latest = 1"
    ).fetchone()["cnt"]

    # Yeni Model sayisi (is_new_model = 1 AND is_latest = 1)
    new_models_cnt = conn.execute(
        "SELECT COUNT(DISTINCT v.model_id) as cnt FROM prices p JOIN variants v ON p.variant_id = v.id WHERE p.is_new_model = 1 AND p.is_latest = 1"
    ).fetchone()["cnt"]

    # Yeni Paket sayisi (is_new_variant = 1 AND is_latest = 1)
    new_variants_cnt = conn.execute(
        "SELECT COUNT(*) as cnt FROM prices WHERE is_new_variant = 1 AND is_latest = 1"
    ).fetchone()["cnt"]

    # Fiyati Degisenler sayisi (price_diff != 0 AND is_latest = 1)
    price_changes_cnt = conn.execute(
        "SELECT COUNT(*) as cnt FROM prices WHERE price_diff != 0 AND price_diff IS NOT NULL AND is_latest = 1"
    ).fetchone()["cnt"]

    # Marka Bazli Ozet Listesi
    brands_summary = conn.execute("""
        SELECT b.name as brand,
               MAX(p.scraped_date) as last_date,
               COUNT(p.id) as total_records,
               SUM(CASE WHEN p.is_latest = 1 THEN 1 ELSE 0 END) as today_records
        FROM brands b
        LEFT JOIN models m ON m.brand_id = b.id
        LEFT JOIN variants v ON v.model_id = m.id
        LEFT JOIN prices p ON p.variant_id = v.id
        GROUP BY b.id, b.name
        ORDER BY b.name
    """).fetchall()

    conn.close()

    return jsonify({
        "last_date": last_date,
        "total_models": total_models,
        "new_models_cnt": new_models_cnt,
        "new_variants_cnt": new_variants_cnt,
        "price_changes_cnt": price_changes_cnt,
        "brands": [dict(b) for b in brands_summary],
        "oem_groups": OEM_GROUPS,
        "scrape_status": SCRAPE_STATUS
    })


@app.route("/api/brands")
@login_required
def api_brands():
    """Aktif markalarin ve araç sayilarinin detayli listesini getirir."""
    conn = get_db()
    rows = conn.execute("""
        SELECT b.name as brand,
               COUNT(DISTINCT m.id) as model_cnt,
               SUM(CASE WHEN p.is_latest = 1 THEN 1 ELSE 0 END) as active_vehicle_cnt
        FROM brands b
        LEFT JOIN models m ON m.brand_id = b.id
        LEFT JOIN variants v ON v.model_id = m.id
        LEFT JOIN prices p ON p.variant_id = v.id
        GROUP BY b.id, b.name
        HAVING active_vehicle_cnt > 0
        ORDER BY b.name ASC
    """).fetchall()
    conn.close()

    tofas_set = set(OEM_GROUPS["tofas"])
    brand_list = []
    for r in rows:
        d = dict(r)
        d["is_tofas_group"] = d["brand"] in tofas_set
        brand_list.append(d)

    return jsonify({
        "brands": brand_list,
        "tofas_brands": OEM_GROUPS["tofas"]
    })


@app.route("/api/models")
@login_required
def api_models():
    brand = request.args.get("brand", "").strip()
    group = request.args.get("group", "").strip()
    conn = get_db()

    query = """
        SELECT DISTINCT m.name as model_name, b.name as brand
        FROM models m
        JOIN brands b ON m.brand_id = b.id
        JOIN variants v ON v.model_id = m.id
        JOIN prices p ON p.variant_id = v.id
        WHERE p.is_latest = 1
    """
    params = []

    if brand and brand.lower() != "all":
        query += " AND LOWER(b.name) = ?"
        params.append(brand.lower())

    if group and group in OEM_GROUPS:
        allowed_brands = [b.lower() for b in OEM_GROUPS[group]]
        query += f" AND LOWER(b.name) IN ({','.join(['?']*len(allowed_brands))})"
        params.extend(allowed_brands)

    query += " ORDER BY m.name ASC"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    return jsonify({"models": [dict(r) for r in rows]})


@app.route("/api/prices")
@login_required
def api_prices():
    group = request.args.get("group", "").strip()
    brand = request.args.get("brand", "").strip()
    model = request.args.get("model", "").strip()
    search = request.args.get("search", "").strip()
    fuel = request.args.get("fuel", "").strip()
    body = request.args.get("body", "").strip()
    trans = request.args.get("trans", "").strip()
    status_filter = request.args.get("status", "").strip()
    year_filter = request.args.get("year", "").strip()
    min_price = request.args.get("min_price", type=int)
    max_price = request.args.get("max_price", type=int)
    only_changes = request.args.get("only_changes", "false").lower() == "true"
    sort_by = request.args.get("sort", "price_asc")

    conn = get_db()

    last_date_row = conn.execute("SELECT MAX(scraped_date) as max_date FROM prices").fetchone()
    target_date = last_date_row["max_date"] if last_date_row else None

    query = """
        SELECT p.id, v.id as variant_id, b.name as brand, m.name as model_name, v.name as variant,
               v.fuel_type, v.transmission, v.engine_power, v.model_year, m.body_type,
               p.price_raw, p.price_int, p.currency, p.scraped_at, p.scraped_date,
               p.is_latest, p.is_new_model, p.is_new_variant, p.previous_price_int,
               p.price_diff, p.price_change_pct
        FROM prices p
        JOIN variants v ON p.variant_id = v.id
        JOIN models m ON v.model_id = m.id
        JOIN brands b ON m.brand_id = b.id
        WHERE 1=1
    """
    params = []

    if request.args.get("all_dates") != "true":
        query += " AND p.is_latest = 1"

    # KURAL: "Tüm Markalar" (group != 'tofas') secili ise Tofaş Grubu markalari ANA LISTEDE GOSTERILMEZ!
    tofas_brands_lower = [b.lower() for b in OEM_GROUPS["tofas"]]
    if group == "tofas":
        query += f" AND LOWER(b.name) IN ({','.join(['?']*len(tofas_brands_lower))})"
        params.extend(tofas_brands_lower)
    elif not brand or brand.lower() == "all":
        # Tüm markalar aktif ve spesifik marka secilmemisse Tofaş markalarini haric tut
        query += f" AND LOWER(b.name) NOT IN ({','.join(['?']*len(tofas_brands_lower))})"
        params.extend(tofas_brands_lower)

    if brand and brand.lower() != "all":
        query += " AND LOWER(b.name) = ?"
        params.append(brand.lower())

    if model and model.lower() != "all":
        query += " AND LOWER(m.name) = ?"
        params.append(model.lower())

    if year_filter and year_filter.lower() != "all":
        query += " AND (v.model_year = ? OR v.name LIKE ?)"
        params.extend([year_filter, f"%{year_filter}%"])

    if fuel and fuel.lower() != "all":
        query += " AND LOWER(v.fuel_type) = ?"
        params.append(fuel.lower())

    if body and body.lower() != "all":
        query += " AND LOWER(m.body_type) LIKE ?"
        params.append(f"%{body.lower()}%")

    if trans and trans.lower() != "all":
        query += " AND LOWER(v.transmission) = ?"
        params.append(trans.lower())

    if min_price:
        query += " AND p.price_int >= ?"
        params.append(min_price)

    if max_price:
        query += " AND p.price_int <= ?"
        params.append(max_price)

    if status_filter == "drop":
        query += " AND p.price_diff < 0"
    elif status_filter == "rise":
        query += " AND p.price_diff > 0"

    if search:
        query += " AND (LOWER(b.name) LIKE ? OR LOWER(m.name) LIKE ? OR LOWER(v.name) LIKE ?)"
        term = f"%{search.lower()}%"
        params.extend([term, term, term])

    if only_changes:
        query += " AND p.price_diff != 0 AND p.price_diff IS NOT NULL"

    # Siralama
    if sort_by == "price_asc":
        query += " ORDER BY p.price_int ASC"
    elif sort_by == "price_desc":
        query += " ORDER BY p.price_int DESC"
    elif sort_by == "diff_asc":
        query += " ORDER BY p.price_diff ASC"
    elif sort_by == "diff_desc":
        query += " ORDER BY p.price_diff DESC"
    else:
        query += " ORDER BY p.price_int ASC"

    rows = conn.execute(query, params).fetchall()
    conn.close()

    result_list = []
    for r in rows:
        d = dict(r)
        d["source_label"] = "Resmi Distribütör Liste Fiyatı"
        result_list.append(d)

    return jsonify({
        "prices": result_list,
        "total": len(result_list),
        "target_date": target_date
    })


@app.route("/api/variant/<int:variant_id>/history")
@login_required
def api_variant_history(variant_id: int):
    """Secilen bir varyantin kapsamli analiz metriklerini ve zaman çizelgesini getir."""
    conn = get_db()
    
    variant_info = conn.execute("""
        SELECT v.id, b.name as brand, m.name as model_name, v.name as variant,
               v.fuel_type, v.transmission, v.engine_power, v.model_year, m.body_type
        FROM variants v
        JOIN models m ON v.model_id = m.id
        JOIN brands b ON m.brand_id = b.id
        WHERE v.id = ?
    """, (variant_id,)).fetchone()

    if not variant_info:
        conn.close()
        return jsonify({"error": "Varyant bulunamadi"}), 404

    history_rows = conn.execute("""
        SELECT id, price_raw, price_int, currency, scraped_at, scraped_date,
               is_latest, previous_price_int, price_diff, price_change_pct
        FROM prices
        WHERE variant_id = ?
        ORDER BY scraped_date ASC, id ASC
    """, (variant_id,)).fetchall()

    conn.close()

    history_list = [dict(h) for h in history_rows]

    # Analiz Metrikleri (KPI Computations)
    start_price = history_list[0]["price_int"] if history_list else 0
    latest_price = history_list[-1]["price_int"] if history_list else 0
    min_price = min([h["price_int"] for h in history_list]) if history_list else 0
    max_price = max([h["price_int"] for h in history_list]) if history_list else 0
    
    net_diff = latest_price - start_price
    net_change_pct = round((net_diff / start_price) * 100, 2) if start_price > 0 else 0.0

    return jsonify({
        "variant": dict(variant_info),
        "analytics": {
            "start_price": start_price,
            "start_date": history_list[0]["scraped_date"] if history_list else "-",
            "latest_price": latest_price,
            "latest_date": history_list[-1]["scraped_date"] if history_list else "-",
            "min_price": min_price,
            "max_price": max_price,
            "net_diff": net_diff,
            "net_change_pct": net_change_pct,
            "total_scans": len(history_list),
        },
        "history": history_list
    })


def _run_scrapers_thread():
    global SCRAPE_STATUS
    SCRAPE_STATUS["running"] = True
    SCRAPE_STATUS["message"] = "Tarama baslatiliyor..."
    SCRAPE_STATUS["progress"] = 0

    conn = sqlite3.connect(str(DB_PATH))
    init_db(conn)

    total_scrapers = len(ALL_SCRAPERS)
    for idx, scraper in enumerate(ALL_SCRAPERS):
        brand_name = scraper.brand
        SCRAPE_STATUS["message"] = f"{brand_name} taraniyor ({idx+1}/{total_scrapers})..."
        SCRAPE_STATUS["progress"] = int(((idx) / total_scrapers) * 100)

        try:
            records, method_used = scraper.run(conn)
            is_stale = method_used == "stale"
            if records and method_used != "failed":
                inserted = save_records(conn, brand_name, records, method_used, 1 if is_stale else 0)
                log_run(conn, brand_name, "fallback" if is_stale else "success", method_used, len(records), f"{inserted} yeni kayit")
            else:
                log_run(conn, brand_name, "error", method_used, 0, "Veri bulunamadi")
        except Exception as exc:
            log_run(conn, brand_name, "error", "exception", 0, str(exc))

    conn.close()
    SCRAPE_STATUS["running"] = False
    SCRAPE_STATUS["message"] = "Tarama tamamlandi!"
    SCRAPE_STATUS["progress"] = 100
    SCRAPE_STATUS["last_run"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


@app.route("/api/trigger-scrape", methods=["POST"])
@login_required
def trigger_scrape():
    global SCRAPE_STATUS
    if session.get("user") != "admin":
        return jsonify({"status": "error", "message": "Canli tarama baslatma yetkisi sadece Admin kullanicisindadir."}), 403

    if SCRAPE_STATUS["running"]:
        return jsonify({"status": "running", "message": "Zaten aktif bir tarama devam ediyor."})

    t = threading.Thread(target=_run_scrapers_thread)
    t.start()
    return jsonify({"status": "started", "message": "Tarama arka planda baslatildi."})


@app.route("/api/export-excel")
@login_required
def api_export_excel():
    """Artan ve azalan fiyatlarin renklendirildigi profesyonel Excel (.xlsx) ihracati."""
    conn = get_db()
    rows = conn.execute("""
        SELECT b.name as brand, m.name as model_name, v.name as variant,
               v.fuel_type, v.transmission, v.engine_power, v.model_year, m.body_type,
               p.price_int, p.currency, p.previous_price_int, p.price_diff, p.price_change_pct, p.scraped_at
        FROM prices p
        JOIN variants v ON p.variant_id = v.id
        JOIN models m ON v.model_id = m.id
        JOIN brands b ON m.brand_id = b.id
        WHERE p.is_latest = 1
        ORDER BY b.name, m.name, p.price_int
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Güncel Araç Fiyatları"
    ws.views.sheetView[0].showGridLines = True

    # Başlık & Hücre Stilleri
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # Renk Dolguları (Fiyatı Artanlar YEŞİL, Düşenler KIRMIZI)
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    green_font = Font(name="Calibri", size=10, bold=True, color="15803D")

    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    red_font = Font(name="Calibri", size=10, bold=True, color="B91C1C")

    data_font = Font(name="Calibri", size=10)
    num_font = Font(name="Calibri", size=10, bold=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    headers = [
        "Marka", "Model", "Varyant / Donanım", "Model Yılı", "Yakit Tipi", "Sanziman", "Motor Gucu",
        "Kasa Tipi", "Guncel Fiyat (TL)", "Onceki Fiyat (TL)", "Fark Tipi", "Fark (TL)", "Degisim (%)", "Guncellenme Tarihi"
    ]

    ws.append(headers)

    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[1].height = 26

    # Verileri Yaz ve Renklendir
    for r_idx, r in enumerate(rows, 2):
        p_diff = r["price_diff"] or 0
        diff_type = "Zam Gelen" if p_diff > 0 else ("Fiyatı Düşen" if p_diff < 0 else "Sabit")

        row_data = [
            r["brand"], r["model_name"], r["variant"], r["model_year"] or "2026",
            r["fuel_type"] or "-", r["transmission"] or "-", r["engine_power"] or "-",
            r["body_type"] or "-", r["price_int"], r["previous_price_int"] or 0,
            diff_type, p_diff, r["price_change_pct"] or 0.0, r["scraped_at"]
        ]
        ws.append(row_data)
        ws.row_dimensions[r_idx].height = 20

        for c_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            cell.font = data_font

            if c_idx in (9, 10, 12):  # Fiyat Sayilari
                cell.number_format = '#,##0 "₺"'
                cell.alignment = align_right
                cell.font = num_font
            elif c_idx == 13:  # % Degisim
                cell.number_format = '+0.00%;-0.00%;0.00%'
                cell.alignment = align_right
            elif c_idx in (4, 5, 6, 7, 8, 11, 14):
                cell.alignment = align_center
            else:
                cell.alignment = align_left

            # Renklendirme Kurallari: Fiyat Artisi YEŞİL, Azalisi KIRMIZI
            if p_diff > 0:
                if c_idx in (9, 11, 12, 13):
                    cell.fill = green_fill
                    cell.font = green_font
            elif p_diff < 0:
                if c_idx in (9, 11, 12, 13):
                    cell.fill = red_fill
                    cell.font = red_font

    # Otomatik Sutun Genisligi
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"arac_fiyat_analizi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/api/export")
@login_required
def api_export():
    return redirect(url_for("api_export_excel"))


if __name__ == "__main__":
    print("=" * 70)
    print("  KURUMSAL COKLU MARKA ARAC FIYAT SCRAPER WEB PANELI")
    print("  Tarayicida acin: http://localhost:5000")
    print("=" * 70)
    app.run(host="0.0.0.0", port=5000, debug=True)
